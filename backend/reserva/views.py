from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from empleados.models import Empleado
from servicio.models import Servicio
from clientes.models import Cliente
from .forms import ReservaForm, ReservaEditForm
from reserva.models import HorarioDisponible, Reserva, PagoReserva, Factura
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta, datetime
from productos.models import Producto
from django.db.models import Count, Q
from django.urls import reverse_lazy
from django.views.generic import CreateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .forms import PagoReservaForm
from django.contrib import messages
from cryptography.fernet import Fernet, InvalidToken
from django.conf import settings
from django.core.mail import send_mail
from django import forms
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.utils.formats import localize
from django.contrib.auth.models import User
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

# Encriptación
fernet = Fernet(settings.ENCRYPT_KEY)

def encrypt_id(pk: int) -> str:
    return fernet.encrypt(str(pk).encode()).decode()

def decrypt_id(token: str) -> int:
    return int(fernet.decrypt(token.encode()).decode())

class ReferenciaPagoForm(forms.Form):
    metodo = forms.ChoiceField(choices=[('nequi', 'Nequi'), ('daviplata', 'Daviplata')], widget=forms.Select(attrs={'class': 'form-select'}))
    referencia = forms.CharField(label='Referencia de pago', max_length=20, widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Número de referencia'}))

def reserva(request):
    gestoras = Empleado.objects.all()
    horarios = HorarioDisponible.objects.all()
    servicios = Servicio.objects.all()
    servicios_por_categoria = {
        'manicure': servicios.filter(categoria='manicure'),
        'pedicure': servicios.filter(categoria='pedicure'),
        'estructura': servicios.filter(categoria='estructura'),
    }

    if request.method == "POST":
        try:
            gestora_id = request.POST.get("gestora")
            nombre_cliente = request.POST.get("nombre")
            telefono_cliente = request.POST.get("telefono")
            servicio_id = request.POST.get("servicio")
            hora_id = request.POST.get("hora")
            fecha_str = request.POST.get("fecha")
            correo_cliente = request.POST.get("correo")

            if not gestora_id or not servicio_id or not hora_id:
                return JsonResponse({"success": False, "error": "Selecciona gestora, servicio y horario."})

            try:
                gestora = Empleado.objects.get(id=gestora_id)
            except Empleado.DoesNotExist:
                return JsonResponse({"success": False, "error": "Gestora no encontrada."})

            try:
                servicio = Servicio.objects.get(id=servicio_id)
            except Servicio.DoesNotExist:
                return JsonResponse({"success": False, "error": "Servicio no encontrado."})

            try:
                hora = HorarioDisponible.objects.get(id=hora_id)
            except HorarioDisponible.DoesNotExist:
                return JsonResponse({"success": False, "error": "Horario no disponible (selecciona otro)."})

            try:
                try:
                    fecha = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                except ValueError:
                    fecha = datetime.fromisoformat(fecha_str).date()
            except Exception:
                return JsonResponse({"success": False, "error": "Formato de fecha inválido."})

            pending = {
                'gestora_id': gestora.id,
                'servicio_id': servicio.id,
                'hora_id': hora.id,
                'fecha': fecha.strftime('%Y-%m-%d'),
                'nombre': nombre_cliente,
                'telefono': telefono_cliente,
                'correo': correo_cliente,
            }

            cliente_obj = Cliente.objects.filter(correo__iexact=correo_cliente).first()
            request.session['pending_reserva'] = pending

            if cliente_obj:
                # Usuario existe, debe iniciar sesión para confirmar reserva
                return JsonResponse({
                    "success": False,
                    "next": "/login/?next=/reserva/completar-reserva/&login_message=Inicia sesión para confirmar tu reserva pendiente."
                })
            else:
                # Usuario no existe, debe registrarse
                return JsonResponse({
                    "success": False,
                    "next": "/login/?register_active=true&next=/reserva/completar-reserva/&pending_message=1&prefill_email={}".format(correo_cliente)
                })

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    if 'pending_reserva' in request.session:
        del request.session['pending_reserva']

    return render(request, 'reservas/reserva.html', {
        'gestoras': gestoras,
        'horarios': horarios,
        'servicios_por_categoria': servicios_por_categoria,
    })

def horarios_disponibles(request):
    fecha = request.GET.get('fecha')
    gestora_id = request.GET.get('gestora_id')
    horarios = HorarioDisponible.objects.all()
    try:
        fecha_obj = datetime.strptime(fecha, '%d/%m/%Y').date()
    except Exception:
        return JsonResponse({'horarios': []})

    ocupados_qs = Reserva.objects.filter(fecha=fecha_obj)
    if gestora_id:
        ocupados_qs = ocupados_qs.filter(gestora_id=gestora_id)
    ocupados = ocupados_qs.values_list('hora_id', flat=True)
    disponibles = horarios.exclude(id__in=ocupados)

    horarios_data = []
    now = timezone.localtime().time()
    es_hoy = fecha_obj == timezone.localdate()
    for h in disponibles:
        hora_str = h.hora.strftime('%H:%M')
        disabled = False
        if es_hoy and h.hora <= now:
            disabled = True
        horarios_data.append({
            'id': h.id,
            'hora': hora_str,
            'disabled': disabled
        })
    return JsonResponse({'horarios': horarios_data})

@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def home(request):
    ahora = timezone.localtime()
    hoy = ahora.date()
    hora_actual = ahora.time().replace(second=0, microsecond=0)

    # Cancelación automática de citas pendientes vencidas
    Reserva.objects.filter(
        estado='pendiente',
        fecha__lt=hoy
    ).exclude(pagos__confirmado=True).update(estado='cancelada')

    Reserva.objects.filter(
        estado='pendiente',
        fecha=hoy,
        hora__hora__lte=hora_actual
    ).exclude(pagos__confirmado=True).update(estado='cancelada')

    # --- AGREGAR ESTE BLOQUE PARA ACTUALIZAR EL ESTADO DESDE LA TABLA ---
    if request.method == "POST" and "reserva_token" in request.POST and "estado" in request.POST:
        try:
            pk = decrypt_id(request.POST["reserva_token"])
            reserva = Reserva.objects.get(pk=pk)
            nuevo_estado = request.POST["estado"]
            if nuevo_estado in dict(reserva.ESTADO_CHOICES):
                reserva.estado = nuevo_estado
                reserva.save()
                messages.success(request, "Estado actualizado correctamente.")
        except Exception as e:
            messages.error(request, "No se pudo actualizar el estado.")
    # --- FIN DEL BLOQUE ---

    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    usuario = request.GET.get('usuario', '')
    estado = request.GET.get('estado', '')
    referencia = request.GET.get('referencia', '').strip()

    citas = Reserva.objects.all()

    if referencia:
        citas = citas.filter(
            Q(pagos__referencia__icontains=referencia) |
            Q(facturas__referencia__icontains=referencia)
        ).distinct()
    else:
        if not estado:
            citas = citas.exclude(estado__in=['cancelada', 'realizada'])
        else:
            citas = citas.filter(estado=estado)

        if fecha_inicio:
            citas = citas.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            citas = citas.filter(fecha__lte=fecha_fin)
        if usuario:
            citas = citas.filter(cliente__nombre__icontains=usuario)

    for cita in citas:
        cita.encrypted_id = encrypt_id(cita.pk)
        
    citas = citas.order_by('fecha', 'hora__hora')

    paginator = Paginator(citas, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'reservas': page_obj,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'usuario': usuario,
        'estado': estado,
        'referencia': referencia,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
        'paginator': paginator,
    }
    return render(request, 'reservas/home.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def agregar_reserva(request):
    fecha = request.GET.get('fecha')
    gestora = request.GET.get('gestora')
    form = ReservaForm(request.POST or None)

    if fecha and gestora:
        ocupadas = Reserva.objects.filter(
            gestora_id=gestora,
            fecha=fecha
        ).values_list('hora_id', flat=True)
        horas_disponibles = HorarioDisponible.objects.exclude(id__in=ocupadas)
        form.fields['hora'].queryset = horas_disponibles.order_by('hora')
    else:
        form.fields['hora'].queryset = HorarioDisponible.objects.all().order_by('hora')

    if request.method == 'POST' and form.is_valid():
        reserva_obj = form.save()
        # Notificación profesional al cliente
        cliente = reserva_obj.cliente
        servicio = reserva_obj.servicio
        fecha_str = reserva_obj.fecha.strftime('%d/%m/%Y')
        hora_str = reserva_obj.hora.hora.strftime('%H:%M')
        empleado = reserva_obj.gestora
        mensaje_html = f"""
        <div style="font-family: 'Montserrat', Arial, sans-serif; background: #f8f9fa; padding: 32px; border-radius: 18px;">
          <h2 style="color: #d63384;">¡Tu cita ha sido agendada!</h2>
          <p>Hola <b>{cliente.nombre}</b>,<br>
          Tu cita para <b>{servicio.nombre}</b> ha sido registrada.<br>
          <b>Fecha:</b> {fecha_str}<br>
          <b>Hora:</b> {hora_str}<br>
          <b>Profesional:</b> {empleado.nombre} {empleado.apellidos}<br>
          </p>
          <hr>
          <p style="color: #d63384;">GiseNails - Tu bienestar, nuestra prioridad.</p>
        </div>
        """
        send_mail(
            subject="Confirmación de cita en GiseNails",
            message="Tu cita ha sido agendada.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[cliente.correo],
            html_message=mensaje_html,
            fail_silently=False,
        )
        return redirect('reserva:home')

    return render(request, 'reservas/agregar.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def editar_reserva(request, token):
    try:
        real_pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Token de reserva inválido.")
        return redirect('reserva:home')
    
    cita = get_object_or_404(Reserva, pk=real_pk)
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in dict(Reserva.ESTADO_CHOICES):
            cita.estado = nuevo_estado
            cita.save()
            # Notificación profesional al cliente
            cliente = cita.cliente
            servicio = cita.servicio
            fecha_str = cita.fecha.strftime('%d/%m/%Y')
            hora_str = cita.hora.hora.strftime('%H:%M')
            estado_display = dict(Reserva.ESTADO_CHOICES).get(nuevo_estado, nuevo_estado)
            mensaje_html = f"""
            <div style="font-family: 'Montserrat', Arial, sans-serif; background: #f9f9fa; padding: 32px; border-radius: 18px;">
              <h2 style="color: #d63384;">Actualización de tu cita</h2>
              <p>Hola <b>{cliente.nombre}</b>,<br>
              El estado de tu cita para <b>{servicio.nombre}</b> ha cambiado a <b>{estado_display}</b>.<br>
              <b>Fecha:</b> {fecha_str}<br>
              <b>Hora:</b> {hora_str}<br>
              </p>
              <hr>
              <p style="color: #d63384;">GiseNails - Siempre a tu servicio.</p>
            </div>
            """
            send_mail(
                subject="Actualización de estado de tu cita en GiseNails",
                message=f"Tu cita ha cambiado a estado {estado_display}.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[cliente.correo],
                html_message=mensaje_html,
                fail_silently=False,
            )
            messages.success(request, "Estado de la cita actualizado correctamente.")
            return redirect('reserva:home')
    else:
        form = ReservaEditForm(instance=cita)
    return render(request, 'reservas/editar.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def eliminar_reserva(request, token):
    try:
        real_pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Token de reserva inválido.")
        return redirect('reserva:home')
    
    cita = get_object_or_404(Reserva, pk=real_pk)
    if request.method == 'POST':
        cita.delete()
        messages.success(request, f"Reserva #{cita.pk} eliminada permanentemente.")
        return redirect("reserva:home")
    return render(request, 'reservas/eliminar.html', {'cita': cita})

class ReservaCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Reserva
    form_class = ReservaForm
    template_name = 'reservas/agregar.html'
    success_url = reverse_lazy('reserva:home')

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        self.object = form.save()
        return super().form_valid(form)

@login_required
def completar_reserva(request):
    pending = request.session.get('pending_reserva')
    if not pending:
        messages.warning(request, "No hay datos de reserva pendientes.")
        return redirect('clientes:panel')

    cliente = Cliente.objects.filter(user=request.user).first()
    if not cliente:
        return redirect('clientes:registro')

    existe = Reserva.objects.filter(
        cliente=cliente,
        gestora_id=pending['gestora_id'],
        servicio_id=pending['servicio_id'],
        hora_id=pending['hora_id'],
        fecha=pending['fecha']
    ).exists()

    if not existe:
        reserva_obj = Reserva.objects.create(
            cliente=cliente,
            gestora=Empleado.objects.get(id=pending['gestora_id']),
            servicio=Servicio.objects.get(id=pending['servicio_id']),
            hora=HorarioDisponible.objects.get(id=pending['hora_id']),
            fecha=pending['fecha'],
            estado='pendiente'
        )
        
        correo_destino = cliente.correo or pending.get('correo')
        if correo_destino:
            servicio_nombre = reserva_obj.servicio.nombre
            fecha_dt = reserva_obj.fecha
            if isinstance(reserva_obj.fecha, str):
                fecha_dt = datetime.fromisoformat(reserva_obj.fecha).date()
            fecha = fecha_dt.strftime('%d/%m/%Y')
            hora = reserva_obj.hora.hora.strftime('%H:%M')
            mensaje = f"Hola {cliente.nombre},\n\nTu cita para '{servicio_nombre}' ha sido agendada para el día {fecha} a las {hora}.\n\n¡Gracias por reservar con Gise-Nails!"
            mensaje_html = f"""
            <div style="font-family: Arial, sans-serif; background: #f9f9f9; padding: 30px;">
                <h2 style="color: #d63384;">¡Cita agendada en Gise-Nails!</h2>
                <p>Hola <b>{cliente.nombre}</b>,</p>
                <p>Tu cita para <b>{servicio_nombre}</b> ha sido agendada para el día <b>{fecha}</b> a las <b>{hora}</b>.</p>
                <p style="margin-top: 40px; color: #555;">¡Gracias por reservar con <b>Gise-Nails</b>!</p>
                <p style="margin-top: 40px;">⚠️ Tu cita debe confirmarse con al menos 1 hora de anticipación. De no hacerlo, será cancelada automáticamente. ¡Gracias por tu puntualidad! </p>
                <h2 style="color: #d63384;">¡Recuerda confirmar tu cita 💖!</h2>
               </div>
            """
            send_mail(
                subject="Confirmación de tu cita en Gise-Nails",
                message=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[correo_destino],
                fail_silently=False,
                html_message=mensaje_html
            )
        
        gestora = reserva_obj.gestora
        servicio = reserva_obj.servicio
        fecha_dt = reserva_obj.fecha
        if isinstance(reserva_obj.fecha, str):
            fecha_dt = datetime.fromisoformat(reserva_obj.fecha).date()
        fecha = fecha_dt.strftime('%d/%m/%Y')
        hora = reserva_obj.hora.hora.strftime('%H:%M')
        mensaje_empleado = (
            f"¡Nueva cita reservada!\n\n"
            f"Cliente: {cliente.nombre}\n"
            f"Teléfono: {cliente.telefono}\n"
            f"Correo: {cliente.correo}\n"
            f"Servicio: {servicio.nombre}\n"
            f"Fecha: {fecha}\n"
            f"Hora: {hora}\n"
        )
        mensaje_html_empleado = f"""
        <div style="font-family: 'Montserrat', Arial, sans-serif; background: #fff0fa; padding: 32px; border-radius: 18px; box-shadow: 0 2px 18px #ffb6e6;">
            <h2 style="color: #d63384; margin-bottom: 12px;">¡Nueva cita agendada!</h2>
            <p style="font-size: 1.15rem; color: #222;">
                <b>Cliente:</b> {cliente.nombre}<br>
                <b>Teléfono:</b> {cliente.telefono}<br>
                <b>Correo:</b> {cliente.correo}<br>
                <b>Servicio:</b> {servicio.nombre}<br>
                <b>Fecha:</b> <span style="color:#d63384;">{fecha}</span><br>
                <b>Hora:</b> <span style="color:#d63384;">{hora}</span>
            </p>
            <hr style="margin: 24px 0; border: none; border-top: 2px solid #ffb6e6;">
            <p style="margin-top: 32px; color: #d63384; font-weight: bold; font-size: 1.2rem;">
                ¡Gracias por tu dedicación! 💅
            </p>
        </div>
        """
        send_mail(
            subject="Nueva cita reservada en Gise-Nails",
            message=mensaje_empleado,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[gestora.correo],
            fail_silently=False,
            html_message=mensaje_html_empleado
        )
    if 'pending_reserva' in request.session:
        del request.session['pending_reserva']

    return redirect('clientes:panel')

@login_required
def confirmar_reserva(request, token):
    try:
        real_pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Enlace de confirmación inválido.")
        return redirect('clientes:panel')

    reserva = get_object_or_404(Reserva, pk=real_pk)
    cliente_email = reserva.cliente.correo
    if not User.objects.filter(username=cliente_email).exists():
        return render(request, 'login/login.html', {
            'register_active': True,
            'prefill_email': cliente_email,
            'pending_message': True
        })
        
    reserva.estado = 'confirmada'
    reserva.save()
    messages.success(request, 'La reserva ha sido confirmada exitosamente.')
    return redirect('clientes:panel')

@login_required
def abonar_reserva(request, token):
    try:
        pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Enlace de reserva inválido.")
        return redirect('clientes:panel')
        
    reserva = get_object_or_404(Reserva, pk=pk, cliente__user=request.user)
    precio_servicio = reserva.servicio.precio if reserva.servicio else Decimal('0')
    monto_abono = precio_servicio * Decimal('0.3')
    monto_abono = int(monto_abono.to_integral_value(rounding='ROUND_HALF_UP'))
    monto_abono_str = localize(monto_abono)

    nequi_num = '3153923380'
    daviplata_num = '3153923380'
    if request.method == 'POST':
        form = ReferenciaPagoForm(request.POST)
        if form.is_valid():
            metodo = form.cleaned_data['metodo']
            referencia = form.cleaned_data['referencia']
            pago = PagoReserva.objects.create(
                reserva=reserva,
                cliente=reserva.cliente,
                monto=monto_abono,
                metodo=metodo,
                confirmado=False,
                referencia=referencia,
                tipo_pago='abono'
            )
            factura, created = Factura.objects.get_or_create(
                reserva=reserva,
                cliente=reserva.cliente,
                defaults={
                    'monto_total': reserva.servicio.precio,  # Guarda el precio actual
                    'pagado': False,
                    'metodo': metodo,
                    'referencia': referencia,
                    'abono': monto_abono,
                    'saldo_restante': reserva.servicio.precio - monto_abono,
                }
            )
            # Si la factura ya existe, NO actualices monto_total si ya hay abono/pago
            if not created and factura.abono > 0:
                # Mantén el monto_total anterior
                pass
            elif not created:
                # Si no hay abono, actualiza el monto_total por si el precio cambió
                factura.monto_total = reserva.servicio.precio
                factura.save()
            # Notificación profesional al cliente
            cliente = reserva.cliente
            servicio = reserva.servicio
            fecha_str = reserva.fecha.strftime('%d/%m/%Y')
            hora_str = reserva.hora.hora.strftime('%H:%M')
            mensaje_html = f"""
            <div style="font-family: 'Montserrat', Arial, sans-serif; background: #f8f9fa; padding: 32px; border-radius: 18px;">
              <h2 style="color: #198754;">¡Pago recibido!</h2>
              <p>Hola <b>{cliente.nombre}</b>,<br>
              Hemos recibido tu abono para el servicio <b>{servicio.nombre}</b>.<br>
              <b>Fecha de la cita:</b> {fecha_str}<br>
              <b>Hora:</b> {hora_str}<br>
              <b>Monto abonado:</b> ${monto_abono}<br>
              <b>Método:</b> {metodo}<br>
              <b>Referencia:</b> {referencia}<br>
              <span style="color: #d63384;">Recuerda que tienes un saldo pendiente de ${factura.saldo_restante}.</span>
              </p>
              <hr>
              <p style="color: #d63384;">GiseNails - Gracias por tu confianza.</p>
            </div>
            """
            send_mail(
                subject="Confirmación de abono en GiseNails",
                message="Hemos recibido tu abono.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[cliente.correo],
                html_message=mensaje_html,
                fail_silently=False,
            )
            messages.success(request, "Referencia enviada. El equipo validará tu pago y confirmará tu cita.")
            return redirect('clientes:panel')
    else:
        form = ReferenciaPagoForm()
    return render(request, 'reserva/abonar_reserva.html', {
        'reserva': reserva,
        'form': form,
        'monto_abono': monto_abono_str,
        'nequi_num': nequi_num,
        'daviplata_num': daviplata_num,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def facturacion(request):
    facturas = Factura.objects.select_related('cliente', 'reserva').order_by('-fecha')
    metodo = request.GET.get('metodo', '')
    estado = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')

    if metodo:
        facturas = facturas.filter(metodo=metodo)
    if estado:
        facturas = facturas.filter(pagado=(estado == 'pagado'))
    if fecha_inicio:
        facturas = facturas.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        facturas = facturas.filter(fecha__lte=fecha_fin)

    return render(request, 'dashboard/facturacion.html', {
        'facturas': facturas,
        'metodo': metodo,
        'estado': estado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })

@login_required
def pagar_saldo(request, token):
    try:
        pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Enlace de pago inválido.")
        return redirect('clientes:panel')
        
    reserva = get_object_or_404(Reserva, pk=pk, cliente__user=request.user)
    factura = reserva.facturas.last()
    
    if not factura:
        messages.error(request, "No existe una factura asociada para pagar el saldo.")
        return redirect('clientes:panel')
        
    monto_saldo = factura.saldo_restante
    form = ReferenciaPagoForm(request.POST or None)
    
    if request.method == 'POST':
        if form.is_valid():
            metodo = form.cleaned_data['metodo']
            referencia = form.cleaned_data['referencia']
            pago = PagoReserva.objects.create(
                reserva=reserva,
                cliente=reserva.cliente,
                monto=monto_saldo,
                metodo=metodo,
                confirmado=False,
                referencia=referencia,
                tipo_pago='saldo'
            )
            factura.saldo_restante = 0
            factura.pagado = True
            factura.metodo = metodo
            factura.referencia = referencia
            factura.save()
            cliente = reserva.cliente
            servicio = reserva.servicio
            fecha_str = reserva.fecha.strftime('%d/%m/%Y')
            hora_str = reserva.hora.hora.strftime('%H:%M')
            mensaje_html = f"""
            <div style="font-family: 'Montserrat', Arial, sans-serif; background: #f8f9fa; padding: 32px; border-radius: 18px;">
              <h2 style="color: #198754;">¡Pago recibido!</h2>
              <p>Hola <b>{cliente.nombre}</b>,<br>
              Hemos recibido tu pago del saldo restante para el servicio <b>{servicio.nombre}</b>.<br>
              <b>Fecha de la cita:</b> {fecha_str}<br>
              <b>Hora:</b> {hora_str}<br>
              <b>Monto pagado:</b> ${monto_saldo}<br>
              <b>Método:</b> {metodo}<br>
              <b>Referencia:</b> {referencia}<br>
              <span style="color: #198754;">¡Tu pago está completo! Nos vemos en tu cita.</span>
              </p>
              <hr>
              <p style="color: #d63384;">GiseNails - Gracias por tu confianza.</p>
            </div>
            """
            send_mail(
                subject="Confirmación de pago de saldo en GiseNails",
                message="Hemos recibido tu pago de saldo.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[cliente.correo],
                html_message=mensaje_html,
                fail_silently=False,
            )
            messages.success(request, "¡Pago completado! Gracias por cancelar el saldo restante.")
            return redirect('clientes:panel')
            
    return render(request, 'reserva/pagar_saldo.html', {
        'reserva': reserva,
        'factura': factura,
        'form': form,
        'monto_saldo': monto_saldo,
    })

@login_required
def pago_efectivo(request, token):
    try:
        pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Enlace de pago inválido.")
        return redirect('clientes:panel')
        
    reserva = get_object_or_404(Reserva, pk=pk, cliente__user=request.user)
    if request.method == 'POST':
        messages.warning(request, "El pago en efectivo se registra en persona. Por favor, realiza el pago en el local.")
        return redirect('clientes:panel')
    return redirect('clientes:panel')

@login_required
def pagar_completo(request, token):
    try:
        pk = decrypt_id(token)
    except InvalidToken:
        messages.error(request, "Enlace de pago inválido.")
        return redirect('clientes:panel')
        
    reserva = get_object_or_404(Reserva, pk=pk, cliente__user=request.user)
    monto_total = reserva.servicio.precio if reserva.servicio else Decimal('0')
    monto_total = int(monto_total.to_integral_value(rounding='ROUND_HALF_UP'))
    monto_total_str = localize(monto_total)
    nequi_num = '3153923380'
    daviplata_num = '3153923380'
    form = ReferenciaPagoForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            metodo = form.cleaned_data['metodo']
            referencia = form.cleaned_data['referencia']
            pago = PagoReserva.objects.create(
                reserva=reserva,
                cliente=reserva.cliente,
                monto=monto_total,
                metodo=metodo,
                confirmado=False,
                referencia=referencia,
                tipo_pago='completo'
            )
            factura, created = Factura.objects.get_or_create(
                reserva=reserva,
                cliente=reserva.cliente,
                defaults={
                    'monto_total': reserva.servicio.precio,
                    'pagado': False,
                    'metodo': metodo,
                    'referencia': referencia,
                    'abono': monto_total,
                    'saldo_restante': 0,
                }
            )
            if not created and factura.abono > 0:
                pass  # Mantén el monto_total anterior
            elif not created:
                factura.monto_total = reserva.servicio.precio
                factura.save()
            cliente = reserva.cliente
            servicio = reserva.servicio
            fecha_str = reserva.fecha.strftime('%d/%m/%Y')
            hora_str = reserva.hora.hora.strftime('%H:%M')
            mensaje_html = f"""
            <div style="font-family: 'Montserrat', Arial, sans-serif; background: #f8f9fa; padding: 32px; border-radius: 18px;">
              <h2 style="color: #198754;">¡Pago recibido!</h2>
              <p>Hola <b>{cliente.nombre}</b>,<br>
              Hemos recibido tu pago completo para el servicio <b>{servicio.nombre}</b>.<br>
              <b>Fecha de la cita:</b> {fecha_str}<br>
              <b>Hora:</b> {hora_str}<br>
              <b>Monto pagado:</b> ${monto_total}<br>
              <b>Método:</b> {metodo}<br>
              <b>Referencia:</b> {referencia}<br>
              <span style="color: #198754;">¡Tu pago está completo! Nos vemos en tu cita.</span>
              </p>
              <hr>
              <p style="color: #d63384;">GiseNails - Gracias por tu confianza.</p>
            </div>
            """
            send_mail(
                subject="Confirmación de pago completo en GiseNails",
                message="Hemos recibido tu pago completo.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[cliente.correo],
                html_message=mensaje_html,
                fail_silently=False,
            )
            messages.success(request, "Referencia enviada. El equipo validará tu pago y confirmará tu cita.")
            return redirect('clientes:panel')
            
    return render(request, 'reserva/pagar_completo.html', {
        'reserva': reserva,
        'form': form,
        'monto_total': monto_total_str,
        'nequi_num': nequi_num,
        'daviplata_num': daviplata_num,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
@csrf_exempt
def pago_efectivo_admin(request, token=None ):
    if request.method == 'POST':
        factura_id = request.POST.get('factura_id')
        referencia = request.POST.get('referencia')
        factura = get_object_or_404(Factura, pk=factura_id, pagado=False)
        reserva = factura.reserva
        cliente = factura.cliente
        monto_pagado = factura.saldo_restante
        PagoReserva.objects.create(
            reserva=reserva,
            cliente=cliente,
            monto=monto_pagado,
            metodo='efectivo',
            confirmado=True,
            referencia=referencia,
            tipo_pago='saldo'
        )
        factura.abono += monto_pagado
        factura.saldo_restante = 0
        factura.pagado = True
        factura.metodo = 'efectivo'
        factura.referencia = referencia
        factura.save()
        cliente = reserva.cliente
        servicio = reserva.servicio
        fecha_str = reserva.fecha.strftime('%d/%m/%Y')
        hora_str = reserva.hora.hora.strftime('%H:%M')
        mensaje_html = f"""
        <div style="font-family: 'Montserrat', Arial, sans-serif; background: #f9f9fa; padding: 32px; border-radius: 18px;">
          <h2 style="color: #198754;">¡Pago recibido!</h2>
          <p>Hola <b>{cliente.nombre}</b>,<br>
          Hemos registrado tu pago en efectivo para el servicio <b>{servicio.nombre}</b>.<br>
          <b>Fecha de la cita:</b> {fecha_str}<br>
          <b>Hora:</b> {hora_str}<br>
          <b>Monto pagado:</b> ${monto_pagado}<br>
          <b>Método:</b> Efectivo<br>
          <b>Referencia:</b> {referencia}<br>
          <span style="color: #198754;">¡Tu pago está completo! Nos vemos en tu cita.</span>
          </p>
          <hr>
          <p style="color: #d63384;">GiseNails - Gracias por tu confianza.</p>
        </div>
        """
        send_mail(
            subject="Confirmación de pago en efectivo en GiseNails",
            message="Hemos registrado tu pago en efectivo.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[cliente.correo],
            html_message=mensaje_html,
            fail_silently=False,
        )
        messages.success(request, "Pago en efectivo registrado correctamente.")
    return redirect('reserva:facturacion')

@login_required
@user_passes_test(lambda u: u.is_staff)
def api_buscar_factura(request):
    ref = request.GET.get('referencia', '').strip()
    factura = Factura.objects.filter(referencia=ref, pagado=False).select_related('cliente', 'reserva').first()
    if factura:
        return JsonResponse({
            'success': True,
            'cliente': factura.cliente.nombre,
            'servicio': str(factura.reserva.servicio),
            'saldo': str(factura.saldo_restante),
            'factura_id': factura.id,
        })
    return JsonResponse({'success': False})

@login_required
@user_passes_test(lambda u: u.is_staff)
def exportar_facturacion_excel(request):
    metodo = request.GET.get('metodo', '')
    estado = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')

    facturas = Factura.objects.all()
    if metodo:
        facturas = facturas.filter(metodo=metodo)
    if estado:
        facturas = facturas.filter(estado=estado)
    if fecha_inicio and fecha_fin:
        facturas = facturas.filter(fecha__range=[fecha_inicio, fecha_fin])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturación"

    # Título y fecha
    ws.merge_cells('A1:I1')
    ws['A1'] = "Reporte de Facturación - GiseNails"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='left')

    # Encabezados
    headers = ["Fecha", "Cliente", "Servicio", "Método", "Referencia", "Pagos", "Saldo", "Total", "Estado"]
    ws.append([])
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Datos
    for factura in facturas:
        pagos = "; ".join([
            f"{pago.get_tipo_pago_display()} - {pago.metodo.title()} - ${pago.monto} - Ref: {pago.referencia}"
            for pago in factura.reserva.pagos.all()
        ])
        ws.append([
            factura.fecha.strftime('%d/%m/%Y %H:%M'),
            factura.cliente.nombre,
            str(factura.reserva.servicio),
            factura.get_metodo_display(),
            factura.referencia,
            pagos,
            factura.saldo_restante,
            factura.monto_total,
            "Pagado" if factura.pagado else "Pendiente"
        ])

    # Ajustar ancho de columnas (evita MergedCell)
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Facturacion_GiseNails.xlsx'
    wb.save(response)
    return response