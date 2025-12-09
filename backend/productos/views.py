import urllib.parse # <-- ¡IMPORTACIÓN CLAVE!
from django.shortcuts import render, redirect, get_object_or_404, Http404
from .models import Producto
from .forms import ProductoForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

# Importar tus funciones de cifrado/descifrado
from django.utils.http import urlsafe_base64_decode
from cryptography.fernet import Fernet
from django.conf import settings

# -----------------------------
# 🔐 Funciones de Cifrado/Descifrado
# -----------------------------

fernet = Fernet(settings.ENCRYPT_KEY)

def encrypt_id(pk):
    return fernet.encrypt(str(pk).encode()).decode()

def decrypt_id(token):
    # Nota: Fernet.decrypt() espera bytes de Base64 seguro para URL (sin %3D)
    return int(fernet.decrypt(token.encode()).decode())


# =====================================================
# 📌 LISTADO
# =====================================================
@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def producto_list(request):
    q = request.GET.get('q', '')
    productos = Producto.objects.all()

    if q:
        productos = productos.filter(nombre__icontains=q)

    paginator = Paginator(productos, 3)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'productos/producto_list.html', {
        'productos': page_obj,
        'q': q,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
        'paginator': paginator,
    })


# =====================================================
# 📌 CREAR
# =====================================================
@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def producto_create(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('producto:producto_list')
    else:
        form = ProductoForm()

    return render(request, 'productos/producto_form.html', {'form': form})


# =====================================================
# 📌 EDITAR (Corregido)
# =====================================================
@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def producto_update(request, token): # <-- Usamos 'token' en la firma
    try:
        # Paso 1: Decodificar URL para limpiar %3D, %2B, etc.
        clean_token = urllib.parse.unquote(token) 
        
        # Paso 2: Desencriptar a ID real
        real_id = decrypt_id(clean_token) 
        producto = get_object_or_404(Producto, pk=real_id)
        
    except Exception:
        # Captura cualquier error de Fernet (InvalidToken) o NotFound
        raise Http404("ID de producto inválido o no encontrado")

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('producto:producto_list')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'productos/producto_form.html', {
        'form': form,
        'producto': producto
    })


# =====================================================
# 📌 ELIMINAR (Corregido)
# =====================================================
@login_required
@user_passes_test(lambda u: u.is_staff)
@never_cache
def producto_delete(request, token): # <-- Usamos 'token' en la firma
    try:
        # Paso 1: Decodificar URL para limpiar %3D, %2B, etc.
        clean_token = urllib.parse.unquote(token) 
        
        # Paso 2: Desencriptar a ID real
        real_id = decrypt_id(clean_token)
        producto = get_object_or_404(Producto, pk=real_id)
        
    except Exception:
        # Captura cualquier error de Fernet (InvalidToken) o NotFound
        raise Http404("ID de producto inválido o no encontrado")

    if request.method == 'POST':
        producto.delete()
        return redirect('producto:producto_list')

    return render(request, 'productos/producto_confirm_delete.html', {
        'producto': producto
    })


# =====================================================
# 📌 DETALLE (Corregido y Limpio)
# =====================================================
def producto_detail(request, token): # <-- Corregido el nombre a solo 'token'
    try:
        # 1. Decodificar URL: Convierte %3D a = y %2F a /
        clean_token = urllib.parse.unquote(token) 
        
        # 2. Desencriptar: Fernet ahora recibe un Base64 válido
        real_id = decrypt_id(clean_token)
        
        producto = get_object_or_404(Producto, pk=real_id)
        
    except Exception:
        # Captura cualquier error de Fernet (InvalidToken) o NotFound
        raise Http404("ID de producto inválido o no encontrado") 

    return render(request, 'productos/producto_detail.html', {
        'producto': producto
    })


# =====================================================
# 📌 LISTA PÚBLICA
# =====================================================
@require_GET
def lista_productos(request):
    productos_list = Producto.objects.all()

    paginator = Paginator(productos_list, 3)
    page_number = request.GET.get('page')
    productos = paginator.get_page(page_number)

    recomendados = Producto.objects.filter(recomendado=True).order_by('-ventas')[:6]
    en_uso = Producto.objects.filter(en_uso=True)[:6]

    print("DEBUG lista_productos: recomendados=", recomendados.count(), " en_uso=", en_uso.count())

    return render(request, 'productos/lista_productos.html', {
        'productos': productos,
        'recomendados': recomendados,
        'en_uso': en_uso,
    })

# =====================================================
# 📌 EXPORTAR A EXCEL
# =====================================================
@login_required
@user_passes_test(lambda u: u.is_staff)
def exportar_productos_excel(request):
    q = request.GET.get('q', '')
    productos = Producto.objects.all()
    if q:
        productos = productos.filter(nombre__icontains=q)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Título y fecha
    ws.merge_cells('A1:E1')
    ws['A1'] = "Reporte de Productos - GiseNails"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='left')

    # Encabezados
    headers = ["ID", "Nombre", "Precio", "Descripción", "Cantidad"]
    ws.append([])
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Datos
    for producto in productos:
        ws.append([
            producto.id,
            producto.nombre,
            producto.precio,
            producto.descripcion,
            producto.cantidad
        ])

    # Ajustar ancho de columnas (evita MergedCell)
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Productos_GiseNails.xlsx'
    wb.save(response)
    return response